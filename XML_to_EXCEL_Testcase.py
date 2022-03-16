# Author : Phan Dong Quang EDA23 - 10/06/2021
from os import name
import xml.etree.ElementTree as ET
import re
import pandas as pd
import openpyxl
from openpyxl import load_workbook 
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter                 
from openpyxl.utils.cell import coordinate_from_string,get_column_interval

class XML_Dict():
        def __init__(self) -> None:
                self.data_dict = {}
                self.test_data_node  = None
                self.tescase_func_gen = {
                        frozenset({'wait','set'}) : self.gen_envvar_wait_func_keyword,
                        frozenset({'wait'}) : self.gen_wait_func_keyword,
                        frozenset({'set'}) : self.gen_envvar_func_keyword,
                        frozenset({'testerconfirmation'}) : self.gen_confirmation_func_keyword,
                        frozenset({'conditions','wait'}) : self.gen_condition_wait_func_keyword

                }

                self.capltestcase_func_data_gen  = {
                        'Wait' :  self.gen_capltest_data_type_2,
                        'SetEnvVar' :  self.gen_capltest_data_type_2,
                        'FunctionalMessage_SPRB' :  self.gen_capltest_data_type_2,
                        'RequestResponse_SPRB' :  self.gen_capltest_data_type_2
                        
                }
  
                self.test_step_tag = ["testcase","capltestcase"]
                self.test_info_tag = "externalref"
                self.test_group_tag = ["testgroup"]
                self.sub_level_key = "sub_test_level"
                self.teststep_level_key = "test_step"

                pass

        def get_XML_data(self,xml_path):
                t_data_list = []
                xml_data = ET.parse(xml_path)
                self.parse_xmL_test_data(xml_data.getroot(),self.update_test_group_data,t_data_list)
                t_data_list = self.gen_cover_node(t_data_list)
                return t_data_list
        
        def gen_cover_node(self,data):
                if len(data) > 1 :
                        result = [{
                                'node' : '',
                                'ID'   : '',
                                "name" : 'Test',
                                self.sub_level_key : data
                        }]
                        
                        return result
                else:
                        return data
                        
        def parse_xmL_test_data(self,xml_node,data_update_func, data_list, level = 1):

                test_sub_tag =  set(list(i.tag for i in xml_node.findall("*")))
                is_test_step_tag = any(list(i in self.test_step_tag for i in test_sub_tag ))
                is_test_group_tag = any(list(i in self.test_group_tag for i in test_sub_tag ))

                sub_elem_ls =  xml_node.findall("./testgroup")
                if len(sub_elem_ls) != 0 or is_test_step_tag:
                
                        if is_test_step_tag or is_test_group_tag :
                                sub_level_data_ls =  data_update_func(xml_node,level,data_list,sub_level_list_key = self.sub_level_key,is_test_step = is_test_step_tag)
                        else:
                                sub_level_data_ls = data_list

                        for i in sub_elem_ls:
                                self.parse_xmL_test_data(i,data_update_func,sub_level_data_ls,level+1)


        def update_test_group_data(self,elem,level,data_list,sub_level_list_key = 'sub_test_level',is_test_step = False):
                if data_list is None :
                        return
                        
                data_dict = self.test_dict_gen(elem,level,sub_level_list_key)
                elem_title =  elem.attrib["title"]
                func_key = None 
                
                if data_dict is not None:
                        data_list.append(data_dict)
                        if is_test_step:
                                step_list = self.update_test_step_data(elem,level)
                                data_list[-1].update({self.teststep_level_key : step_list })

                        return data_dict[sub_level_list_key]
                else:
                        return data_list

        def test_dict_gen(self,xml_node,level_name ,sub_level_list_key = None):
                if xml_node is not None:
                        _ID = xml_node.find("./externalref")
                        if _ID is not None:
                                ID = _ID.attrib["title"] 
                        else :
                                # ID = " "
                                return  

                        name  =  xml_node.attrib["title"]
                else:
                     pass

                result = {
                        'node'		                : xml_node,
                        'ID'   		                : ID,
                        'name' 		                : name,
                        'level'                         : level_name,
                }

                if sub_level_list_key is not None:
                        result.update({ f'{sub_level_list_key}': list()})

                return result

        def update_test_step_dict(self,test_step_node,test_level):
                '''
                        update test step require information
                '''
                if test_step_node.tag == "externalref":
                        return None
                result =  {
                                        'step'		: self.gen_test_step_str(test_step_node),
                                        'response'	: self.gen_test_response_str(test_step_node),
                                        'keyword'	: self.gen_test_key_work_str(test_step_node),
                                        'level'         : test_level
                }
        
                return result

        def update_test_step_data(self,test_group_node,level,data_list = None):
                # step_node_list = test_group_node.findall("*")
                data_list = [] if data_list is None else data_list
                step_node_list = list(i for i in test_group_node.findall("*") if i.tag in self.test_step_tag)

                for step_node in step_node_list:
                        step = self.update_test_step_dict(step_node,level)
                        if step is not None:
                                data_list.append(step)
                return data_list

        
        # driver test step update func
        def gen_test_step_str(self,test_step_node):
                result = "-"
                atrib_key =  "title"
                if atrib_key  in test_step_node.attrib:
                        result =  self.trim_index_num(test_step_node.attrib[atrib_key ])
                return result

        def gen_test_response_str(self,test_step_node):
                result = "-"
                atrib_key = 'ident'
                if atrib_key in test_step_node.attrib:
                        result =  self.trim_index_num(test_step_node.attrib[atrib_key])
                return result

        def gen_test_key_work_str(self,test_step_node):
                keyword = ''
                if 'testcase' == test_step_node.tag :
                        keyword = self.gen_testcase_tag_keyword(test_step_node)
                elif 'capltestcase' == test_step_node.tag:
                        keyword  = self.gen_capltestcase_tag_keyword(test_step_node)
                return keyword
        
        def trim_index_num(self,string):
                tmp = re.findall(r'(^[\d,\),\-,\\,\/]+)',string)
                if len(tmp) != 0 :
                        result = re.findall(r'^[\d,\),\-,\\,\/]+\s(.*)',string)
                        return string if  len(result) == 0 else result[0]
                else:
                        return string
                        
        # split test step case func
        def gen_testcase_tag_keyword(self,test_step_node):
                # sub_node_list = list(filter(lambda x: len(x) == 0,list(test_step_node.iter('*'))))
                sub_node_list = test_step_node.findall("./*")
                sub_node_tag = frozenset(list(node.tag for node in sub_node_list))
                
                if sub_node_tag is not None:
                        keyword = self.tescase_func_gen.get(sub_node_tag)(test_step_node)
                else :
                        print(20*"-------")
                        print(f"There is no keyword gen function support for xml tag  {list(node.tag for node in sub_node_list)}")
                        keyword =  "CAN NOT GEN KEYWORD"

                return keyword

        def gen_capltestcase_tag_keyword(self,test_step_node):
                '''
                        Gen string format for capltestcase tag node
                        Assumption :  only one command is used in each capltestcase
                '''
                test_step_node_name = test_step_node.attrib["name"]
                get_data_func = self.capltestcase_func_data_gen.get(test_step_node_name,self.gen_default_capltest_data)

                caplparam_node = test_step_node.findall("./caplparam")
                data_list = [get_data_func(node) for node in caplparam_node]
                name = test_step_node.attrib['name']
                
                keyword =  self.gen_keyword_string(name,data_list,separator= ",")
                        
                return keyword


        # function to gen specific command from xml format
        def gen_envvar_wait_func_keyword(self,test_step_node):
                command = ''
                sub_command = []
                sub_node_list =  list(test_step_node.iter('*'))[1:]

                # get list of leaf node in test_step_node
                # assumption : envvar and wait tag come in pair
                envvar_node_list =  [node for node in sub_node_list if len(node) == 0 and node.tag == 'envvar']
                wait_node_list =  [node for node in sub_node_list if len(node) == 0 and node.tag == 'wait']


                for envvar,wait in zip(envvar_node_list,wait_node_list):
                        name =  envvar.attrib['name']	
                        data_list = [envvar.text,wait.attrib['time']]
                        sub_command.append(self.gen_keyword_string(name,data_list,separator=';'))
                
                # gen final string
                # format : tag_name(subnode command)
                command =  self.gen_keyword_string('envvar',sub_command,separator=',')

                return command

        def gen_envvar_func_keyword(self,test_step_node):
                sub_node = test_step_node.find('.//envvar')
                data = sub_node.attrib['name']
                sub_keyword = self.gen_keyword_string(data,sub_node.text,separator = '') 
                keyword = self.gen_keyword_string(sub_node.tag,list(sub_keyword),separator = '')
                return keyword

        def gen_wait_func_keyword(self,test_step_node):
                sub_node_wait = test_step_node.find('./wait')
                wait_time = sub_node_wait.attrib['time']
                keyword = self.gen_keyword_string(sub_node_wait.tag,list(wait_time),separator = '')

                return keyword
        
        def gen_confirmation_func_keyword(self,test_step_node):
                sub_node = test_step_node.find('./testerconfirmation')
                data = sub_node.attrib['title']
                keyword = self.gen_keyword_string(sub_node.tag,list(data),separator = '')
                return keyword

        def gen_keyword_string(self,name,data_list,separator = ',' ):
                for i in range(len(data_list)):
                        if data_list[i] is None:
                                data_list[i] = ''

                data_string = f"{separator}".join(x for x in data_list)
                return f"{str(name)}({data_string})"

        def gen_capltest_data_type_2(self,step_node):
                name  =  step_node.attrib["name"]
                type  =  step_node.attrib["type"]
                text  =  step_node.text
                data_keyword =  " ".join([type,name,text])

                return data_keyword

        def gen_default_capltest_data(self,step_node):
                return step_node.text

        def gen_condition_wait_func_keyword(self,test_step_node):
                func_list =  {
                                        'dlc_ok' : self.gen_dlc_ok_funckeyword,
                                        'cycletime_rel' : self.gen_cycletime_func_keyword,
                }

                node_attrib_title  = list(node.tag for node in test_step_node.findall('.//*'))
                keyword_gen_func = None
                for key,data in func_list.items():
                        if key in node_attrib_title:
                                keyword_gen_func = data
                                break
                        
                keyword = keyword_gen_func(test_step_node) if keyword_gen_func is not None else  "CAN NOT GEN KEYWORD"
                
                return keyword
        
        
        def gen_cycletime_func_keyword(self,test_step_node):
                cycletime_node = test_step_node.find('.//cycletime_rel')
                min_time = cycletime_node.attrib["min"]
                max_time = cycletime_node.attrib["max"]
                sub_node =cycletime_node.find('./canmsg')
                sig_name = sub_node.attrib['id']
                bus_name= sub_node.attrib['bus']
                data = [sig_name,bus_name,min_time,max_time]
                keyword = self.gen_keyword_string('cycletime',data,separator = ',')
                return keyword

        def gen_dlc_ok_funckeyword(self,test_step_node):
                canmsg_node = test_step_node.find('.//canmsg')
                node_id =  canmsg_node.attrib["id"]
                node_bus =  canmsg_node.attrib["bus"]
                data = [node_id,node_bus]
                keyword = self.gen_keyword_string('dlc_ok',data,separator = ',')
                return keyword     

class CSV_gen():
        def __init__(self) -> None:
                self.header = ["ID",
                                "MDC DCOM Tests",
                                "Test Description",
                                "TestSteps",
                                "Test response",
                                "Teststep keywords",
                                "ObjectType",
                                "Project",
                                ]
                self.leaf_node_key = ""
                self.sub_level_key = ""
                self.data_dict = dict.fromkeys(self.header,None)
        
        def set_leaf_node_key(self,key):
                self.leaf_node_key = key
                pass

        def set_sub_level_key(self,key):
                self.sub_level_key = key
                pass

        def gen_report(self,test_data,excel_path = "excel_xml.xlsx",project_name = " "):
                self.parse_test_data(test_data,self.sub_level_key,self.leaf_node_key)
                self.update_project_name(project_name)
                test_df = pd.DataFrame(self.data_dict)
                print('Write excel data - Done')
                test_df.to_excel(excel_path,index= False, freeze_panes= (0,0))
                print('Save excel data - Done')
                pass

        def update_project_name(self,project_name):
                for i in range(len(self.data_dict[self.header[-1]])):
                        self.data_dict[ self.header[-1]][i] = project_name
                pass

        def parse_test_data(self,test_data,sub_level_key_ls,leaf_node_key,level = None):
                for index,elem in enumerate(test_data, start= 1):

                        index = f"{level}.{index}" if level is not None else f"{index}"  
                        
                        if leaf_node_key in elem.keys():
                                self.update_test_case_data(elem,index)
                        
                        elif sub_level_key_ls in elem.keys():
                                self.update_test_group_data(elem,index)
                               
                        sub_level_test_data = elem[sub_level_key_ls]
                        if len(sub_level_test_data) != 0:
                                self.parse_test_data(sub_level_test_data,sub_level_key_ls,leaf_node_key,index)

                pass

        def fill_background(self,excel_path,name_list,back_ground_colour):
                # find row that contain value of name list
                wb =  load_workbook(excel_path)
                ws =  wb[wb.sheetnames[0]]
                col_index = self.header.index("MDC_DCOM_Tests") + 1
                col_str =  get_column_letter(col_index)
                
                for cell in ws[col_str]:
                        tmp = any([x in cell.value for x in name_list])
                        if tmp:
                                # for i in range(1,len(self.header)+1):
                                cell.fill = PatternFill( start_color= "dbf2de",fill_type = "solid")
                wb.save(excel_path)
                # fill the background with properly colour

                pass
        
        def update_test_group_data(self,data_dict,index):
                name = f'{index} ' + data_dict['name']
                # data ={
                #         'MDC DCOM Tests' : name ,
                #         'ID' : data_dict['ID'],
                #         "ObjectType" : 'TestGroup'

                # }
                data ={
                        # 'MDC DCOM Tests'
                        self.header[1] : name , 
                        # 'ID'
                        self.header[0] : data_dict['ID'],
                        # "ObjectType"
                        self.header[6] : 'TestGroup'

                }

                self.update_data_dict(data,self.data_dict)
                pass

        def update_test_case_data(self,case_dict,index):
                test_step_list = list(self.gen_dict_extract('test_step',case_dict))[0]
                step_str =  self.gen_test_step(test_step_list)

                object_type = "Automated Testcase" if  any(len(i) for i in step_str) else "TestGroup"                      

                # data ={
                #         'MDC DCOM Tests' : f'{index} ' + case_dict['name'],
                #         'ID' : case_dict['ID'],
                #         "ObjectType" : object_type,
                #         "TestSteps": step_str[0],
                #         "Test response":  step_str[1],
                #         "Teststep keywords": step_str[2]
                # }

                data ={
                        #  'MDC DCOM Tests' 
                        self.header[1]: f'{index} ' + case_dict['name'],
                        #  'ID' 
                        self.header[0]: case_dict['ID'],
                        #  "ObjectType" 
                        self.header[6]: object_type,
                        #  "TestSteps"
                        self.header[3]: step_str[0],
                        #  "Test response"
                        self.header[4]:  step_str[1],
                        #  "Teststep keywords"
                        self.header[5]: step_str[2]
                }
                
                self.update_data_dict(data,self.data_dict)

        def update_data_dict(self,update_data,dict):
                for key,value in list(self.data_dict.items()) :
                        
                        if key in update_data:
                                update_value = update_data[f'{key}']
                        else:
                                update_value = ''

                        if value is None:
                                dict.update({f'{key}' : [update_value]})
                        else:
                                value.append(update_value)

        def gen_test_step(self,test_step_list):
                step_str = None
                response_str = None
                keyword_str = None
                
                step = []
                response = []
                keyword = []

                for test_step in test_step_list:
                        step.extend(list(self.gen_dict_extract('step',test_step)))
                        response.extend(list(self.gen_dict_extract('response',test_step)))
                        keyword.extend(list(self.gen_dict_extract('keyword',test_step)))

                step_str = self.list_to_string(step)
                response_str = self.list_to_string(response)
                keyword_str = self.list_to_string(keyword)
        
                return step_str,response_str,keyword_str

        def list_to_string(self,step_list):
                result =''
                for i,step in enumerate(step_list,start= 1):
                        result += ''.join([f'{i})',step,'\n'])
                return result

        def gen_dict_extract(self,key_value, var,key_return = None):
                '''
                        extract value base on key in all nested dict
                '''
                # if hasattr(var,'items'):
                for k, v in var.items():
                        if type(key_value) is dict and key_return is not None:
                                if {k:v} == key_value:
                                        yield var.get(key_return,None)

                        elif k == key_value:
                                yield v
                        
                        if isinstance(v, dict):
                                for result in self.gen_dict_extract(key_value, v,key_return):
                                        yield result
                        elif isinstance(v, list):
                                for d in v:
                                        for result in self.gen_dict_extract(key_value, d,key_return):
                                                yield result
        


if __name__ == "__main__":
        xml = XML_Dict()
        data = xml.get_XML_data(f"C:/Users/PUQ81HC/Documents/MPC/Tool/XML_excel/Tool_test_data/7/TC_gen.xml")
        csv = CSV_gen()
        csv.set_leaf_node_key(xml.teststep_level_key)
        csv.set_sub_level_key(xml.sub_level_key)
        csv.gen_report(data,excel_path = f"C:/Users/PUQ81HC/Documents/MPC/Tool/XML_excel/Tool_test_data/7/TC_gen.xlsx",project_name= 'abc')
